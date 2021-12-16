import openpyxl
from openpyxl import *

from tqdm import tqdm

# excel 파일들을 변수에 로드
excel_tagging = load_workbook("result_data_star.xlsx")
excel_today_news_keyword = load_workbook("result_data.xlsx")

sheet_model_word = excel_tagging.active
sheet_today_news_keyword = excel_today_news_keyword.active

# 값들을 집어넣을 변수
model_keyword_data = {}
today_keyword_data = {}
value = 0


print("loading keyword")

# 날짜별로 키워드 전체 저장
for i in tqdm(range(1, sheet_model_word.max_row+1)):
    model_keyword_data[sheet_model_word.cell(row=i, column=1).value] = sheet_model_word.cell(row=i, column=2).value
for i in tqdm(range(1, sheet_today_news_keyword.max_row+1)):
    today_keyword_data[sheet_today_news_keyword.cell(row=i, column=1).value] = sheet_today_news_keyword.cell(row=i, column=2).value
print("loading data")

for i in tqdm(range(1, sheet_today_news_keyword.max_row+1)):
    if sheet_model_word.cell(row=i, column=1).value == sheet_today_news_keyword.cell(row=i, column=2).value:
        value = value + today_keyword_data[sheet_today_news_keyword.cell(row=i, column=1).value]
        
print(value)
print("program ended")