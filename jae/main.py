from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import gensim

from tqdm import tqdm

from time import sleep

# 사전 훈련된 Word2Vec 모델을 로드합니다.
model = gensim.models.KeyedVectors.load_word2vec_format('model1.bin', binary=True)

wb1 = load_workbook("model_word2.xlsx")
wb2 = load_workbook("뉴스전체키워드재정렬.xlsx")
wb3 = load_workbook("samsung_finance_data.xlsx")
ws1 = wb1.active
ws2 = wb2.active
ws3 = wb3.active

# w2v 해라
w2v_word = ['' for i in range(300000000)]  # w2v 처리 끝난 단어들
listday1 = ['' for i in range(10000000)]
listday2 = ['' for i in range(10000000)]
keywords = {}
stocks_open = {}
stocks_close = {}

print("program started")

print(ws2.max_row)
sleep(10000)

for y in tqdm(range(1, ws2.max_column + 1)):
    listday1.append(ws2.cell(row=y, column=1).value)
for y in tqdm(range(1, ws3.max_column + 1)):
    listday2.append(ws3.cell(row=y, column=1).value)

for i in tqdm(range(1, ws2.max_column + 1)):
    keywords[listday1[i]] = ws2.cell(row=y, column=2).value.split(',')
for y in tqdm(range(1, ws3.max_column + 1)):
    stocks_open[listday2[i]] = ws3.cell(row=y, column=2).value
    stocks_close[listday2[i]] = ws3.cell(row=y, column=3).value

print(keywords[listday1[i]])


sleep(1000)


for i in range(1, len(listday1)+2):
    for k in tqdm(range(1, len(listday1)+2)):
        try:
            w2v_word.append(model.most_similar(keywords[listday1[k]])[0:2])  # 여기에 워드투 벡터 처리한 단어 넣는다. 날짜별로 키워드는 keywords[listday[i]]로 불러올 수 있다. 인덱스는 1부터 시작해야함
        except:
            None
    for y in range(1, ws1.max_column + 1):
        for j in range(1, len(w2v_word) + 2):
            if ws1.cell(row=y, column=1).value == w2v_word[j][0]:
                if stocks_open[listday2[i]] > stocks_close[listday2[i]]:
                    ws1.cell(row=y, column=2).value = int(ws1.cell(row=y, column=2).value) + float(w2v_word[j][0])
                elif stocks_open[listday2[i]] < stocks_close[listday2[i]]:
                    ws1.cell(row=2, column=y).value = int(ws1.cell(row=2, column=y).value) - float(w2v_word[j][0])

wb1.save("result2222.xlsx")

print("program ended")