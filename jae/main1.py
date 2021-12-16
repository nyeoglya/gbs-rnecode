import openpyxl
from openpyxl import *

import gensim

from tqdm import tqdm

from time import sleep

print("loading word2vec model...")

# 사전 훈련된 Word2Vec 모델을 로드
model = gensim.models.KeyedVectors.load_word2vec_format('model1.bin', binary=True)

# excel 파일들을 변수에 로드
excel_model_word = load_workbook("model_word2.xlsx")
excel_news_keyword = load_workbook("news_keyword.xlsx")
excel_finance = load_workbook("samsung_finance_data.xlsx")

sheet_model_word = excel_model_word.active
sheet_news_keyword = excel_news_keyword.active
sheet_finance = excel_finance.active

# 값들을 집어넣을 변수
data = {}
delta_finance = {}
model_keyword_data = {}

date_list = ['' for i in range(sheet_finance.max_row)]

print("loading keyword")

# 날짜별로 키워드 전체 저장
for i in tqdm(range(1, sheet_news_keyword.max_row+1)):
    data[sheet_news_keyword.cell(row=i, column=1).value] = sheet_news_keyword.cell(row=i, column=2).value

print("loading finance data")

# 날짜별로 주식 증감을 측정해 기록
for i in tqdm(range(1, sheet_finance.max_row+1)):
    try:
        if len(sheet_finance.cell(row=i, column=1).value) > 6:
            date_list[i-1] = sheet_finance.cell(row=i, column=1).value
        if int(sheet_finance.cell(row=i, column=2).value) <= int(sheet_finance.cell(row=i, column=3).value):
            delta_finance[sheet_finance.cell(row=i, column=1).value] = 1
        else:
            delta_finance[sheet_finance.cell(row=i, column=1).value] = -1
    except:
        None

print("loading keyword_data")

# 뉴스 키워드 데이터 전체를 저장
for i in tqdm(range(1, sheet_model_word.max_row+1)):
    model_keyword_data[sheet_model_word.cell(row=i, column=1).value] = 0

# 미리 만들어진 Word2Vec 데이터를 활용해 연산을 수행
print("calculating")
for i in tqdm(range(0,8)): # tqdm(range(0, sheet_news_keyword.max_row)):
    temp_data = data[date_list[i]].split(',')
    cnt = len(temp_data)
    for j in range(0, cnt):
        try:
            for k in range(0, 2):
                temp_data2 = model.most_similar(temp_data[j])
                model_keyword_data[temp_data2[k][0]] += temp_data2[k][1] * delta_finance[date_list[i]]
        except:
            # print("에러 발생")
            None

# 완성된 dictionary를 리스트로 변경
save_data_list = list(model_keyword_data.items())

print("saving data")

# 모든 데이터를 다시 집어넣음
excel_save = openpyxl.Workbook()
sheet_save = excel_save.active

# 데이터를 저장하는 코드
for i in range(0, len(model_keyword_data)+5): # 수정하기
    try:
        sheet_save.cell(row=i, column=1).value = str(save_data_list[i][0])
        sheet_save.cell(row=i, column=2).value = str(save_data_list[i][1])
    except:
        None

excel_save.save("result_data_star.xlsx")

print("program ended")

