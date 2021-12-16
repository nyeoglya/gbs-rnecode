# 파일의 개수와 파일명을 읽기 위한 모듈
import os

#  액셀 파일을 열기 위한 모듈
import openpyxl
from openpyxl import *

# 진행 상황을 알기 위한 프로세스 바를 출력하는 모듈
from tqdm import tqdm

# sleep 함수를 사용하기 위한 time 모듈
from time import sleep

line_count = 0 # 모든 데이터의 줄 개수

cnt = 0 # 임시 변수
k = 0 # 임시 변수2

# 수용 가능한 데이터 최대값
max_data = 10000

date_data = ['' for i in range(max_data)] # 데이터별 날짜
data = ['0' for i in range(max_data)] # 모든 데이터를 저장할 변수
temp_date_data = ['' for i in range(250)]
temp_data = ['' for i in range(250)]

path = "./news_data/" # 액셀 파일이 있는 경로
file_list = os.listdir(path) # 경로 안에 있는 파일명을 불러모음

# news_data 디렉토리내의 모든 파일명 출력
print("\n\nfound files: " + ",".join(file_list) + "\n")

file_count = len(file_list) # news_data 내부에 있는 파일의 개수를 출력

# 파일 개수에 대해 다음을 실행
for i in range(0, file_count):
    print("now reading: " + file_list[i]) # 어떤 파일을 읽는지에 대해 출력
    excel = load_workbook(path + file_list[i]) # 액셀 파일을 excel이라는 변수 안에 로딩
    
    sheet = excel.active
    
    # 액셀 파일 내부의 줄의 개수를 cnt에 저장
    cnt = 0
    k = 0
    stack = "None" # 임시 변수 stack을 만들고 이를 이용해 값들을 정렬
    for j in sheet.rows:
        cnt += 1
        line_date = str(j[1].value)
        
        if str(j[18].value).find("중복") != -1: continue
        elif str(j[18].value).find("예외") != -1: continue
        
        if stack == line_date:
            temp_data[k] = temp_data[k] + j[15].value + ',' + j[14].value
        else:
            if line_date != "일자":
                stack = line_date
                k += 1
                temp_date_data[k] = line_date
                
                print("now reading: " + line_date)
            else:
                k += 1
    
    
    line_count += k
    j = 0
    
    # 데이터 전체 역정렬
    print("데이터 역정렬중\n")
    for j in range(0, k):
        date_data[line_count-k+j] = temp_date_data[k-j]
        data[line_count-k+j] = temp_data[k-j]
        
        # print(date_data[line_count-k+j])
    
    print("\n데이터 역정렬 완료됨\n")
    
    # cnt를 출력함으로 액셀 파일 내부의 줄의 개수를 파악함
    print("value counted: " + str(cnt))
    
    print("\n=======================\n") # 다음 작업과 구분을 위한 선

# 모든 데이터를 다시 집어넣음
excel_save = openpyxl.Workbook()
sheet_save = excel_save.active

for j in range(1, line_count):
    print(date_data[j] + ":" + str(j))
    
    sheet_save.cell(row=j, column=1).value = date_data[j]
    sheet_save.cell(row=j, column=2).value = data[j]

excel_save.save("result_data.xlsx")

print("\n\ndata saved!\nlines: " + str(line_count) + "\n\n")
