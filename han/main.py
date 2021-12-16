import openpyxl
from openpyxl import *

from tqdm import tqdm

from time import sleep
"""
import pandas as pd
import matplotlib as plt
import numpy as np
"""

date = []
title = []
news_keyword = []
raw_keywords = []
length_data = 0
twined_data = []
twined_date = []
max_data = 23670

def func_openpy():
    start_year = 2017
    start_month = 1
    end_year = 2017
    end_month = 4
    news_list = ["NewsResult_20170119-20170419.xlsx","NewsResult_20170419-20170719.xlsx","NewsResult_20170719-20171019.xlsx",
    "NewsResult_20171019-20180119.xlsx","NewsResult_20180119-20180419.xlsx", "NewsResult_20180419-20180719.xlsx",
    "NewsResult_20180719-20181019.xlsx","NewsResult_20181019-20190119.xlsx","NewsResult_20190119-20190419.xlsx",
    "NewsResult_20190419-20190719.xlsx","NewsResult_20190719-20191019.xlsx","NewsResult_20191019-20200119.xlsx",
    "NewsResult_20200119-20200419.xlsx","NewsResult_20200419-20200719.xlsx","NewsResult_20200719-20201019.xlsx",
    "NewsResult_20201019-20210119.xlsx","NewsResult_20210119-20210419.xlsx","NewsResult_20210419-20210719.xlsx",
    "NewsResult_20210719-20211019.xlsx"
    ]

    for news in range(0,1): # (0,19)
        global length_data
        global max_data
        j = 2

        excel = load_workbook("news_data/" + news_list[news])

        sheet = excel.active
        for j in tqdm(range(2,max_data+1)):
            cell = sheet.cell(row=j,column=2)
            if cell.value is not None:
                date.append(cell.value)
        
        for j in tqdm(range(2,max_data+1)):
            cell = sheet.cell(row=j,column=15)
            raw_keywords = cell.value
            
            if cell.value is not None:
                try:
                    news_keyword.append(raw_keywords.split(','))
                except:
                    None
                    
        print(type(excel))
        print("reading succesfuly ended")
        excel.close()

"""
def func_pandas():
    df = pd.read_excel('news_data/NewsResult_20170119-20170418.xlsx',)
"""

def integrate_data():
    stack = "NA"
    n = -1
    
    date_length = len(date)
    print(date_length)
    
    for i in range(0,date_length):
        if (stack == date[i]): # 스택 값에 들어있는 값이 날짜 값과 같다면
            try:
                for j in range(0,len(news_keyword[i])):
                    try:
                        twined_data[n][1].append(str(news_keyword[i][j])) # 날짜 값의 데이터를 같은 날짜 값의 twined_data로 옮긴다
                    except:
                        print(news_keyword[i])
            except:
                None
        else: # 아니라면
            n += 1 # 다음 날짜로 옮김
            stack = str(date[i])  # 스택 값을 변경
            
            twined_data.append([[],[]])
            
            try:
                twined_data[n][0].append(str(date[i])) # 날짜 값을 설정
            except:
                print(date[i] + ":" + str(n))
                quit()
    
    """
    try:
        for i in range(0,date_length):
            if(stack == date[i]):
                for j in range(0,len(news_keyword[i])):
                    twined_data[n][1].append(news_keyword[i][j])
            else:
                n += 1
                stack = date[i]
                #twined_date.append(stack)
                twined_data.append([[],[]])
                twined_data[n][0] = date[i]
    """

def print_bundle():
    print(str(len(date)) + ":" + str(len(news_keyword)))
    print(date)
    sleep(1000)
    print("----------------------------")
    print(news_keyword)
    print("----------------------------")
    print(twined_data)


def data_re():
    excel = load_workbook('result_data.xlsx')
    sheet = excel.active
    word_bundle = ''

    for i in tqdm(range(1,len(twined_data)+1)):
        if "".join(twined_data[i-1][0]) == "None": break
        
        print("".join(twined_data[i-1][0]))
        
        sheet.cell(row=i,column=1).value = "".join(twined_data[i-1][0])
        word_bundle = word_bundle + ','.join(twined_data[i-1][1])
        sheet.cell(row=i,column=2).value = word_bundle
        word_bundle = ''

    excel.save('result_data.xlsx')

def main():
    func_openpy()
    
    integrate_data()
    # print_bundle()
    
    data_re()
    print("program succesfuly ended")

main()