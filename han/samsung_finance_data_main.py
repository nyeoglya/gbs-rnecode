#  액셀 파일을 열기 위한 모듈
import openpyxl
from openpyxl import *

line_count = 0 # 모든 데이터의 줄 개수

# 수용 가능한 데이터 최대값
max_data = 2000

date_data = ['' for i in range(max_data)] # 데이터별 날짜

excel = load_workbook("samsung_finance_data.xlsx")
sheet = excel.active

cnt = 0
for row in sheet.rows:
    date_data[cnt] = str(row[0].value).replace('-', '').split(' ')[0]
    cnt += 1


# 모든 데이터를 다시 집어넣음
excel_save = openpyxl.Workbook()
sheet_save = excel_save.active

for j in range(1, cnt+1):
    sheet_save.cell(row=j, column=1).value = date_data[j]

excel_save.save("samsung_finance_result.xlsx")
